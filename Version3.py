import yfinance as yf
import pandas as pd
import time
import tkinter as tk
from tkinter import messagebox, simpledialog
from tkinter import ttk
from tkcalendar import DateEntry
from threading import Thread
from yahooquery import search, Ticker
from openpyxl import load_workbook

# Global flag to stop fetching process
cancel_flag = False


# Function to fetch data and save to Excel
def fetch_and_save_data(tickers, start_date, end_date, interval, filename, progress_var, progress_label):
    global cancel_flag
    total_tickers = len(tickers)

    # Create DataFrames to store individual stock data
    individual_data = {}

    # Create DataFrames to store combined data
    combined_data = {stat: pd.DataFrame() for stat in ['Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume']}

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for i, ticker in enumerate(tickers):
            if cancel_flag:
                messagebox.showinfo("Cancelled", "Data fetching process was cancelled.")
                break
            try:
                print(f"Fetching data for {ticker} with interval {interval}")
                stock_data = yf.download(ticker, start=start_date, end=end_date, interval=interval)

                # Ensure the index is datetime
                stock_data.index = pd.to_datetime(stock_data.index)

                # Write individual stock data to sheets
                stock_data[['Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume']].to_excel(writer, sheet_name=ticker)

                # Store individual stock data
                individual_data[ticker] = stock_data

                # Combine individual data for each statistic
                for stat in ['Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume']:
                    combined_data[stat][ticker] = stock_data[stat]

                time.sleep(2)  # Delay to avoid hitting rate limits
            except Exception as e:
                error_message = f"Error fetching data for {ticker}: {e}"
                print(error_message)
                messagebox.showerror("Error", error_message)

            # Update progress
            progress = int((i + 1) / total_tickers * 100)
            progress_var.set(progress)
            progress_label.config(text=f'Progress: {progress}%')

        # Write combined data for each statistic to separate sheets, in the desired order
        for stat in ['Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume']:
            combined_data[stat].to_excel(writer, sheet_name=stat)

    # Re-open the file to adjust date format and column width
    wb = load_workbook(filename)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        for cell in ws['A']:
            if cell.row == 1:
                continue
            cell.number_format = 'yyyy-mm-dd'
        ws.column_dimensions['A'].width = 70 / 7  # Set column width to 70 pixels

    wb.save(filename)

    if not cancel_flag:
        messagebox.showinfo("Success", "Data saved successfully to " + filename)

    # Reset cancel flag
    cancel_flag = False


# Function to handle the fetch button click
def on_fetch_click():
    global cancel_flag
    cancel_flag = False
    tickers = ticker_entry.get().split(',')
    start_date = start_date_entry.get_date().strftime('%Y-%m-%d')
    end_date = end_date_entry.get_date().strftime('%Y-%m-%d')
    interval = interval_var.get()
    filename = simpledialog.askstring("Input", "Enter the filename for the Excel file:",
                                      initialvalue="historical_prices.xlsx")
    if filename:
        # Run data fetching in a separate thread to keep the GUI responsive
        thread = Thread(target=fetch_and_save_data,
                        args=(tickers, start_date, end_date, interval, filename, progress_var, progress_label))
        thread.start()


# Function to handle the cancel button click
def on_cancel_click():
    global cancel_flag
    cancel_flag = True


# Function to search for ticker symbol and display company names
def search_ticker():
    query = search_entry.get()
    if query:
        try:
            search_results = search(query)
            if search_results and 'quotes' in search_results:
                matches = [(item['symbol'], item['shortname']) for item in search_results['quotes'] if
                           'symbol' in item and 'shortname' in item]
                if matches:
                    # Display the results in a new window
                    result_window = tk.Toplevel(root)
                    result_window.title("Search Results")
                    tk.Label(result_window, text="Select a ticker symbol:").pack(pady=10)

                    result_var = tk.StringVar(value=matches[0][0])
                    for symbol, name in matches:
                        tk.Radiobutton(result_window, text=f"{symbol} - {name}", variable=result_var,
                                       value=symbol).pack(anchor='w')

                    def select_ticker():
                        ticker_entry.insert(tk.END, result_var.get() + ',')
                        result_window.destroy()

                    tk.Button(result_window, text="Select", command=select_ticker).pack(pady=10)
                else:
                    messagebox.showinfo("No Results", "No ticker symbols found for the given query.")
            else:
                messagebox.showinfo("No Results", "No ticker symbols found for the given query.")
        except Exception as e:
            messagebox.showerror("Error", "Error fetching data: " + str(e))


# Create the main window
root = tk.Tk()
root.title("Stock Data Fetcher")

# Customize calendar style
style = ttk.Style(root)
style.theme_use('clam')
style.configure('TCombobox', selectbackground='white', selectforeground='black')  # Combobox style
style.configure('TButton', background='lightgray')  # Button style
style.configure('TLabel', background='lightgray', foreground='black')  # Label style
style.configure('TEntry', fieldbackground='white')  # Entry style
style.configure('Calendar.Treeview', background='white', foreground='black', fieldbackground='white')  # Calendar style

# Create and place the widgets
tk.Label(root, text="Search Company:").grid(row=0, column=0, padx=10, pady=10)
search_entry = tk.Entry(root, width=50)
search_entry.grid(row=0, column=1, padx=10, pady=10)

search_button = tk.Button(root, text="Search Ticker", command=search_ticker)
search_button.grid(row=0, column=2, padx=10, pady=10)

tk.Label(root, text="Tickers (comma separated):").grid(row=1, column=0, padx=10, pady=10)
ticker_entry = tk.Entry(root, width=50)
ticker_entry.grid(row=1, column=1, padx=10, pady=10)

tk.Label(root, text="Start Date:").grid(row=2, column=0, padx=10, pady=10)
start_date_entry = DateEntry(root, width=12, background='darkblue', foreground='white', borderwidth=2, year_bg='white',
                             month_fg='black')
start_date_entry.grid(row=2, column=1, padx=10, pady=10)

tk.Label(root, text="End Date:").grid(row=3, column=0, padx=10, pady=10)
end_date_entry = DateEntry(root, width=12, background='darkblue', foreground='white', borderwidth=2, year_bg='white',
                           month_fg='black')
end_date_entry.grid(row=3, column=1, padx=10, pady=10)

tk.Label(root, text="Interval:").grid(row=4, column=0, padx=10, pady=10)
interval_var = tk.StringVar()
interval_combobox = ttk.Combobox(root, width=10, textvariable=interval_var, state='readonly')
interval_combobox['values'] = ('1d', '1wk', '1mo')
interval_combobox.current(0)
interval_combobox.grid(row=4, column=1, padx=10, pady=10)

fetch_button = tk.Button(root, text="Fetch Data", command=on_fetch_click)
fetch_button.grid(row=5, column=0, padx=10, pady=10)

cancel_button = tk.Button(root, text="Cancel", command=on_cancel_click)
cancel_button.grid(row=5, column=1, padx=10, pady=10)

progress_var = tk.IntVar()
progress_label = tk.Label(root, text='Progress: 0%')
progress_label.grid(row=6, column=0, columnspan=2, padx=10, pady=10)

root.mainloop()
